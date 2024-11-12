"""
Pontificia Universidad Javeriana
Autor    : Carlos Eduardo Rojas
Fecha    : 29/10/2024
Materia  : Sistemas Operativos
Tema     : Análisis de Rendimiento de Ejecuciones Concurrentes
Fichero  : Procesamiento y Exportación de Datos de Ejecución
Objetivo : Procesar y analizar datos de rendimiento en archivos
           de ejecución concurrente, calculando promedios y 
           exportando los resultados a un archivo Excel.
"""


import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Ruta de la carpeta que contiene los archivos de datos
folder_path = 'C:\\Users\\carlo\\OneDrive - Pontificia Universidad Javeriana\\Desktop\\Uni\\SistemasOperativos\\MaquinasProyecto\\DatosEjecucion'
output_file_path = 'C:\\Users\\carlo\\OneDrive - Pontificia Universidad Javeriana\\Desktop\\Uni\\SistemasOperativos\\MaquinasProyecto\\Datos.xlsx'

# Lista para almacenar todos los datos procesados
data_list = []

# Obtener y ordenar la lista de archivos
files = sorted(os.listdir(folder_path), key=lambda x: (int(re.search(r'-(\d+)-Hilos', x).group(1)), int(re.search(r'Hilos-(\d+)', x).group(1))))

# Recorre cada archivo en la carpeta DatosEjecucion
for filename in files:
    print(f"Procesando archivo: {filename}")  # Mensaje de depuración
    # Verifica que sea un archivo .txt o .dat
    if filename.endswith('.txt') or filename.endswith('.dat'):
        # Extrae el nombre del archivo y genera el encabezado
        #match = re.match(r'MM_ejecutable-(\d+)-Hilos-(\d+)', filename)
        match = re.match(r'mc.exe-(\d+)-Hilos-(\d+)', filename)
        if match:
            hilos = match.group(1)
            cores = match.group(2)
            header = f"{hilos} Hilos {cores} Core{'s' if cores != '1' else ''}"
            print(f"Encabezado generado: {header}")  # Mensaje de depuración
            
            # Lee el contenido del archivo actual
            file_path = os.path.join(folder_path, filename)
            with open(file_path, 'r') as input_file:
                # Procesa cada línea y almacena los datos limpios en una lista
                values = []
                for line in input_file:
                    data = re.search(r'(\d+)', line)
                    if data:
                        values.append(int(data.group(1)))
            
            # Añadir los datos a la lista con el encabezado correspondiente
            # Se agregarán los valores originales y sus promedios en bloques de 30
            values_with_averages = []
            for i in range(0, len(values), 30):
                chunk = values[i:i+30]  # Subconjunto de 30 valores
                values_with_averages.extend(chunk)  # Añade los 30 valores
                if len(chunk) == 30:  # Si hay exactamente 30 valores
                    avg = sum(chunk) / 30
                    values_with_averages.append(avg)  # Añade el promedio
                
            data_list.append((header, values_with_averages))

# Crear un diccionario para convertir a DataFrame
data_dict = {header: pd.Series(values) for header, values in data_list}

# Crear el DataFrame y guardarlo como un archivo Excel
df = pd.DataFrame(data_dict)
df.to_excel(output_file_path, index=False)

# Ajustar el ancho de las columnas y añadir color de fondo a las celdas de los encabezados y promedios
workbook = load_workbook(output_file_path)
worksheet = workbook.active

# Ajustar el ancho de las columnas de A hasta BL a 15.89 (equivalente a 150 píxeles aproximadamente)
for col in worksheet.columns:
    worksheet.column_dimensions[col[0].column_letter].width = 15.89

# Crear un relleno de color para las celdas de los títulos y promedios
title_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
average_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Aplicar el relleno de color a las celdas de los encabezados
for cell in worksheet[1]:  # Solo afecta la primera fila (encabezados)
    cell.fill = title_fill

# Aplicar el relleno de color a las celdas de promedios
for col in worksheet.columns:
    for cell in col:
        if isinstance(cell.value, float):  # Si la celda contiene un promedio
            cell.fill = average_fill

workbook.save(output_file_path)
print("Archivo 'Datos.xlsx' generado con éxito con promedios y color de encabezado y promedios ajustados.")
